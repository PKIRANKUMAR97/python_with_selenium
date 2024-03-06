from faker import Faker
from openpyxl import Workbook

fake_data = Faker(['hi_IN','en_US'])
wb = Workbook()
ws = wb.active
for i in range(1, 11):
    for j in range(1, 5):
        ws.cell(row=i, column=j, value=fake_data.name())
        '''
        ws.cell(row=i, column=2, value=fake_data.email())
        ws.cell(row=i, column=3, value=fake_data.address())
        ws.cell(row=i, column=4, value=fake_data.phone_number())
        '''
wb.save("test_data.xlsx")

'''
print(fake_data.name())
print(fake_data.email())
print(fake_data.address())
'''
