from openpyxl import Workbook,load_workbook
wb = load_workbook(filename='demoexcel.xlsx')
sh = wb['DemoSheet']
row_c=sh.max_row
col_c=sh.max_column
for i in range(1,row_c+1):
    for j in range(1,col_c+1):
        print(sh.cell(row=i,column=j).value,end=' ')
    print('\n')
#print(row_c)
#print(col_c)
#print(sh['A2'].value)