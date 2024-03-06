from openpyxl import Workbook

wb = Workbook()
ws = wb.active
# ws['C8']="KIRAN"
'''
testdata = [['NAME', 'CITY'], ['KIRAN', 'VIZAG'],['RAZZ','DELHII']]
for data in testdata:
    ws.append(data)
'''
for i in range(1,6):
    for j in range(1,5):
        ws.cell(row=i,column=j).value=i+j
wb.save("demoexcel_1.xlsx")
